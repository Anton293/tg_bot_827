"""/get_week, get_rosp, edit into /set [value]"""
from telegram import InlineKeyboardButton, InlineKeyboardMarkup
from openpyxl import Workbook, load_workbook
import os
import datetime
import sys

#data command
from root.data_users import data
dirname = os.path.basename(os.path.dirname(__file__))
storage_command = data.data_commands[dirname]
from root.default import admin_command

print(storage_command['version'])


keyboard_week = [[
    InlineKeyboardButton("Понеділок", callback_data="week_0"),
    InlineKeyboardButton("Вівторок", callback_data="week_1"),
    InlineKeyboardButton("Середа", callback_data="week_2"),
], [
    InlineKeyboardButton("Четверг", callback_data="week_3"),
    InlineKeyboardButton("Пятниця", callback_data="week_4"),
], [
    InlineKeyboardButton("Подробней", callback_data="week_details")
]]
reply_markup_week = InlineKeyboardMarkup(keyboard_week)

keyboard_today = [
    #InlineKeyboardButton("Редактировать", callback_data="week_edit"),
    InlineKeyboardButton("назад", callback_data="week_down")
]
reply_markup_today = InlineKeyboardMarkup(keyboard_today)
keyboard_edition = [
    InlineKeyboardButton("Зберегти", callback_data="week_save"),
    InlineKeyboardButton("Відмінити", callback_data="week_cancel"),
    InlineKeyboardButton("назад", callback_data="week_down")
]


########################################################################
#                       CREATED DAY CONTENT                            #
########################################################################
wb = load_workbook("res/db/schedule/database.xlsx")
ws = wb.active
data.arr_days_week = ['Понеділок', "Вівторок", "Середа", "Четверг", "Пятниця"]
data.arr_time_couple = ['08:00=>9:00', '09:00=>10:00', '10:00=>9:00', '11:00=>13:00']


def time_couple_read_from_table():
    def get_time_couple(i):
        try:
            return f"{ws[f'K{i + 2}'].value.strftime('%H:%M')} => {ws[f'L{i + 2}'].value.strftime('%H:%M')}"
        except AttributeError:
            return None
    return [get_time_couple(i) for i in range(9) if get_time_couple(i) is not None]


data.arr_time_couple = time_couple_read_from_table()
print(data.arr_time_couple)


def content_c(letter, next_row):
    value = ws[f'{letter}{next_row}'].value
    if value is None:
        return "Нет пари"
    return value



def context_head():
    day_id = data.TODAY
    bd_in_text = ""
    week_number = datetime.datetime.today().isocalendar()[1]
    type_week = 0
    choice_couple = []
    array_week = ws.iter_rows(min_row=2, max_row=50)
    for i, row in enumerate(array_week):
        if row[0].value is not None and row[6].value is not None and i not in choice_couple:
            if row[0].value - 1 == day_id and row[6].value == 1:
                couple_id = row[1].value - 1
                bd_in_text += f"\n\n*{couple_id+1}◷{data.arr_time_couple[couple_id]}*"
                next_row = i+3
                if ws[f"G{next_row}"].value == 1 and row[1].value == ws[f"B{next_row}"].value:
                    choice_couple.append(i+1)

                    bd_in_text += f"\n\t\t\t\t\t\t\t> Ч: {row[2].value}"
                    bd_in_text += f"\n\t\t\t\t\t\t\t> З: {content_c('C', next_row)}"
                elif row[6].value is None or row[6].value == 0 or row[2].value is None:
                    bd_in_text += f"\n\t\t\t\t\t\t\t> Нет пари!"
                else:
                    bd_in_text += f"\n\t\t\t\t\t\t\t> {row[2].value}"

    return f"Расписание на: ({week_number}){data.arr_days_week[day_id]}\t{bd_in_text}"


########################################################################
#                              CALL WEEK                               #
########################################################################


@admin_command
def get_button_options(update, _):
    """update message options for view information week"""
    update.message.reply_text(context_head(), parse_mode="Markdown", reply_markup=reply_markup_week)


def see_week(query):
    """view the week command /get_week"""
    print("see week")
    query.edit_message_text(context_head(), parse_mode="Markdown", reply_markup=reply_markup_week)


########################################################################
#                                DETAILS                               #
########################################################################


def created_button_couple(day_id):
    choice_couple = []
    arr_button = []
    array_week = ws.iter_rows(min_row=2, max_row=50)
    for i, row in enumerate(array_week):
        if row[0].value is not None and row[6].value is not None and i not in choice_couple:
            if row[0].value - 1 == day_id and row[6].value == 1:
                next_row = i + 3
                if ws[f"G{next_row}"].value == 1 and row[1].value == ws[f"B{next_row}"].value:
                    choice_couple.append(i + 1)
                    arr_button.append(InlineKeyboardButton(f"Ч:{row[1].value}", callback_data=f"week_couple_{i}"))
                    arr_button.append(InlineKeyboardButton(f"З:{content_c('B', next_row)}", callback_data=f"week_couple_{i + 1}"))
                else:
                    arr_button.append(InlineKeyboardButton(f"{row[1].value}", callback_data=f"week_couple_{i}"))
    return arr_button


def details(query):
    print("details")
    day_id = data.TODAY
    couple_id = data.couple

    arr_button = created_button_couple(day_id)
    keyboard = [arr_button, keyboard_today]
    reply_markup_local = InlineKeyboardMarkup(keyboard)

    query.edit_message_text(f"""*{data.arr_days_week[day_id]}* > {ws[f'B{couple_id + 2}'].value} пара\n
    ◉**({couple_id+1}) {content_c("C", couple_id+2)}**
    ◉*Час: *{data.arr_time_couple[ws[f'B{couple_id + 2}'].value - 1]}
    ◉*Платформа:* {ws[f'F{couple_id+2}'].value}
    ◉*Викладач:* {ws[f'E{couple_id+2}'].value}
        \n◉*Посилання:* {ws[f'D{couple_id+2}'].value}
    """, reply_markup=reply_markup_local, parse_mode="Markdown")


########################################################################
#                                  BRAIN                               #
########################################################################


def processing_keyboard_week(update, query, query_key):
    query.data = query_key
    if query.data == "down":
        see_week(query)
        return
    elif query.data == "details":
        details(query)
        return
    elif "couple" in query.data and not int(query.data.split('_')[1]) == data.couple:
        print("couple")
        data.couple = int(query.data.split('_')[1])
        details(query)
        return

    if not data.TODAY == int(query_key):
        data.TODAY = int(query_key)
        see_week(query)
        return
    return
