from openpyxl import Workbook, load_workbook

wb = load_workbook("res/root/database.xlsx")
ws = wb.active
print(ws['A1'].value)


